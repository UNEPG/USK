import os
import sys
import string
import openpyxl
from copy import copy
from os import listdir
from pathlib import Path
from os.path import isfile, join
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

column_list = []

def main_merge():
    global column_list
    input_folder = "mrg_input"
    dir_path = os.path.dirname(os.path.realpath(__file__)) + "\\" + input_folder
    filename = Path(os.path.abspath(__file__)).resolve().stem + '.py'
    dir_files = [f for f in listdir(dir_path) if isfile(join(dir_path, f))]
    dir_files_count = len(dir_files)

    print("\r\nBOM files merger. {}".format(filename))

    max_column = 0
    min_column = 100

    for x in range(0, dir_files_count):
        input_table = load_workbook(filename = input_folder + "\\" + dir_files[x])
        input_sheet = input_table.active
        if input_sheet.max_column > max_column:
            max_column = input_sheet.max_column
        if input_sheet.max_column < min_column:
            min_column = input_sheet.max_column

    if min_column != max_column:
        print("Attention! Column count not equal")
        pass #sys.exit()

    # rows = reader(dir_files[0], input_folder + "\\")
    # for row in rows:
    #     output_sheet.append(row)
    base_filename = dir_files[0]
    output_table = load_workbook(filename = input_folder + "\\" + base_filename)
    dir_files.remove(dir_files[0])
    dir_files_count = len(dir_files)
    #dir_files_count = 2
    column_list = list(string.ascii_uppercase)
    column_list = list(map(chr, range(ord('A'), ord('Z')+1)))


    output_filename = 'mrg_output/Merged_list.xlsx'

    output_sheet = output_table.active
    base_file_title = output_sheet.title
    output_sheet.title = "Merged BOM"
    output_isheet = output_table.create_sheet(title="Information")
    output_isheet['A1'] = "Input files"
    output_isheet['A2'] = base_filename
    output_isheet['D2'] = count_row(output_sheet, 1)

    for x in range(0, dir_files_count):
        output_isheet['A{}'.format(x+3)] = dir_files[x]

    designator_column = 0
    quantity_column = 0
    man_pn_column = 0

    for x in range(1, max_column+1):
        if output_sheet.cell(column=x, row=1)._value and \
            'Quantity' in output_sheet.cell(column=x, row=1)._value:
            quantity_column = x
            break
    for x in range(1, max_column+1):
        if output_sheet.cell(column=x, row=1)._value and \
            'Manufacturer Part Number' in output_sheet.cell(column=x, row=1)._value:
            man_pn_column = x
            break
    for x in range(1, max_column+1):
        if output_sheet.cell(column=x, row=1)._value and \
            'Designator' in output_sheet.cell(column=x, row=1)._value:
            designator_column = x
            break

    output_sheet_row_count = count_row(output_sheet, 1)
    for x in range(2, output_sheet_row_count+1):
        output_sheet.cell(column=designator_column, row=x)._value = '{}( {} )'.format(base_file_title, output_sheet.cell(column=designator_column, row=x)._value)
        
    for x in range(0, dir_files_count):
        input_table = load_workbook(filename = input_folder + "\\" + dir_files[x])
        input_sheet = input_table.worksheets[0]
        input_sheet_row_count = count_row(input_sheet, 1)

        output_isheet['D{}'.format(x+3)] = input_sheet_row_count - 1
        
        for y in range(2, input_sheet_row_count + 1):
            found = False
            output_sheet_row_count = count_row(output_sheet, 1)
            for z in range(2, output_sheet_row_count + 1):
                input_man_pn = input_sheet.cell(column=man_pn_column, row=y)._value
                output_man_pn = output_sheet.cell(column=man_pn_column, row=z)._value
                if input_man_pn == None or output_man_pn == None:
                    continue
                if type(input_man_pn) is not str:
                    input_man_pn = str(input_man_pn)
                if type(output_man_pn) is not str:
                    output_man_pn = str(output_man_pn)
                if len(input_man_pn) < 3 and len(output_man_pn) < 3:
                    continue

                if input_man_pn in output_man_pn or output_man_pn in input_man_pn:
                    output_sheet.cell(column=quantity_column, row=z)._value = output_sheet.cell(column=quantity_column, row=z)._value + \
                                                                                    input_sheet.cell(column=quantity_column, row=y)._value
                    
                    output_sheet.cell(column=designator_column, row=z)._value = output_sheet.cell(column=designator_column, row=z)._value + ', \n' \
                                                                                            + '{}( {} )'.format(input_sheet.title, input_sheet.cell(column=designator_column, row=y)._value)
                    found = True
                    break

            if not found:
                output_sheet_row_count += 1
                output_sheet.insert_rows(output_sheet_row_count)
                copy_cells( output_sheet, output_sheet_row_count, input_sheet, y, max_column )
                output_sheet.cell(column=designator_column, row=output_sheet_row_count)._value = \
                                        '{}( {} )'.format(input_sheet.title, output_sheet.cell(column=designator_column, row=output_sheet_row_count)._value)

    output_sheet_row_count = count_row(output_sheet, 1)
    output_table.save(filename = output_filename)
    print("\r\n")


def copy_cells(target, target_row, source, source_row, max, no_style=False, no_link=False):
    for x in range(1, max+1):
        target_cell = target.cell(column=x, row=target_row)
        source_cell = source.cell(column=x, row=source_row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if no_style and source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if no_link and source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

def count_row( sheet, column ):
    sheet_row_count = 0
    while True:
        if sheet.cell(column=1, row=sheet_row_count+1)._value == None:
            break
        sheet_row_count += 1
    return sheet_row_count

def compare( output, folder, filename_A, filename_B, quantity_column=3, man_pn_column=6, max_column=8 ):

    input_sheet_A = load_workbook(filename = folder + "\\" + filename_A).active
    input_sheet_A_row_count = count_row(input_sheet_A, 1)
    input_sheet_B = load_workbook(filename = folder + "\\" + filename_B).active
    input_sheet_B_row_count = count_row(input_sheet_B, 1)

    for x in range(2, input_sheet_A_row_count + 1):
        
        found = False
        need_copy = False
        
        for y in range(2, input_sheet_B_row_count + 1):
            input_sheet_A_man_pn = input_sheet_A.cell(column=man_pn_column, row=x)._value
            input_sheet_B_man_pn = input_sheet_B.cell(column=man_pn_column, row=y)._value
            if input_sheet_A_man_pn == None or input_sheet_B_man_pn == None:
                continue
            if type(input_sheet_A_man_pn) is not str:
                input_sheet_A_man_pn = str(input_sheet_A_man_pn)
            if type(input_sheet_B_man_pn) is not str:
                input_sheet_B_man_pn = str(input_sheet_B_man_pn)
            if len(input_sheet_A_man_pn) < 3 and len(input_sheet_B_man_pn) < 3:
                continue

            if input_sheet_A_man_pn in input_sheet_B_man_pn or input_sheet_B_man_pn in input_sheet_A_man_pn:
                found = True
                dif_quantity = 0
                input_sheet_A_quantity = input_sheet_A.cell(column=quantity_column, row=x)._value
                input_sheet_B_quantity = input_sheet_B.cell(column=quantity_column, row=y)._value
                if input_sheet_A_quantity == input_sheet_B_quantity:
                    need_copy = False
                    #break
                else:
                    
                    dif_quantity = input_sheet_A_quantity - input_sheet_B_quantity

                    input_sheet_A.cell(column=quantity_column, row=x)._value = dif_quantity

                    need_copy = True
                    #break

        if need_copy or not found:
            output_sheet_row_count = count_row(output, 1) + 1
            output.insert_rows(output_sheet_row_count)
            copy_cells( output, output_sheet_row_count, input_sheet_A, x, max_column )

    return output

def main_compare(both_dir=False):
    global column_list
    input_folder = "cmp_input"
    filename = Path(os.path.abspath(__file__)).resolve().stem + '.py'
    dir_path = os.path.dirname(os.path.realpath(__file__)) + "\\" + input_folder
    dir_files = [f for f in listdir(dir_path) if isfile(join(dir_path, f))]
    dir_files_count = len(dir_files)

    if dir_files_count != 2:
        sys.exit()

    print("\r\nBOM files comparing. {}".format(filename))

    max_column = 0
    min_column = 100

    for x in range(0, dir_files_count):
        input_table = load_workbook(filename = input_folder + "\\" + dir_files[x])
        input_sheet = input_table.active
        if input_sheet.max_column > max_column:
            max_column = input_sheet.max_column
        if input_sheet.max_column < min_column:
            min_column = input_sheet.max_column

    if min_column != max_column:
        print("Attention! Column count not equal")
        pass #sys.exit()

    output_table = load_workbook(filename = input_folder + "\\" + dir_files[0])
    output_sheet = output_table.active
    output_sheet_row_count = count_row(output_sheet, 1)
    output_sheet.delete_rows(2, output_sheet_row_count - 1 )
    output_sheet.title = "Compared list"
    output_filename = 'cmp_output/Compared_list.xlsx'

    designator_column = 0
    quantity_column = 0
    man_pn_column = 0

    for x in range(1, max_column+1):
        if output_sheet.cell(column=x, row=1)._value and \
            'Quantity' in output_sheet.cell(column=x, row=1)._value:
            quantity_column = x
            break
    for x in range(1, max_column+1):
        if output_sheet.cell(column=x, row=1)._value and \
            'Manufacturer Part Number' in output_sheet.cell(column=x, row=1)._value:
            man_pn_column = x
            break
    for x in range(1, max_column+1):
        if output_sheet.cell(column=x, row=1)._value and \
            'Designator' in output_sheet.cell(column=x, row=1)._value:
            designator_column = x
            break

    output_sheet = compare(output_sheet, input_folder, dir_files[0], dir_files[1])
    if both_dir:
        output_sheet = compare(output_sheet, input_folder, dir_files[1], dir_files[0])
    
    output_sheet_row_count = count_row(output_sheet, 1)
    output_table.save(filename = output_filename)
    print("\r\n")

if len(sys.argv) > 1 and sys.argv[1] == '--compare':
    if len(sys.argv) > 2 and sys.argv[2] == '--both_dir':
        main_compare(both_dir=True)
    else:
        main_compare()
else:
    main_merge()
